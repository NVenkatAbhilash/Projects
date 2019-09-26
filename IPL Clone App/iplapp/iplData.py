import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'iplleague.settings')
django.setup()
import openpyxl
from iplapp.models import *
import csv

def dumpMatches():
    matches_excel = openpyxl.load_workbook('resource/matches.xlsx')
    matches_sheet = matches_excel.get_sheet_by_name('Worksheet')
    rows = matches_sheet.max_row
    for row in range(2, rows+1):
        match = Matches()
        match.id = matches_sheet.cell(row=row, column=1).value
        match.season = matches_sheet.cell(row=row, column=2).value
        match.city = matches_sheet.cell(row=row, column=3).value
        date = matches_sheet.cell(row=row, column=4).value
        if '/' in date:
            day, month, year = date.split('/')
            date = '20'+year+'-'+month+'-'+day
        match.date = date
        match.team1 = matches_sheet.cell(row=row, column=5).value
        match.team2 = matches_sheet.cell(row=row, column=6).value
        match.toss_winner = matches_sheet.cell(row=row, column=7).value
        match.toss_decision = matches_sheet.cell(row=row, column=8).value
        match.result = matches_sheet.cell(row=row, column=9).value
        match.dl_applied = matches_sheet.cell(row=row, column=10).value
        match.winner = matches_sheet.cell(row=row, column=11).value
        match.win_by_runs = matches_sheet.cell(row=row, column=12).value
        match.win_by_wickets = matches_sheet.cell(row=row, column=13).value
        match.player_of_match = matches_sheet.cell(row=row, column=14).value
        match.venue = matches_sheet.cell(row=row, column=15).value
        match.umpire1 = matches_sheet.cell(row=row, column=16).value
        match.umpire2 = matches_sheet.cell(row=row, column=17).value
        match.umpire3 = matches_sheet.cell(row=row, column=18).value
        match.save()


def dumpDeliveries():
    os.environ.setdefault('DJANGO_SETTINGS_MODULE',"ipl_league.settings")
    django.setup()
    with open('resource/deliveries.csv','rt') as f:
        data = csv.DictReader(f)
        for row in data:
            match = Deliveries(**row)
            match.save()


if __name__ == '__main__':
    dumpMatches()
    dumpDeliveries()